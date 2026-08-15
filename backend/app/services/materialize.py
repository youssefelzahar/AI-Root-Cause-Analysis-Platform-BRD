"""Convert source files into a canonical typed Parquet rendering.

Every source type - CSV, Excel, SQL Server - converges on Parquet, so the
profiler and the future RCA engine have exactly one reader and never need to
know where the data came from (PRD section 11).
"""

import os
import tempfile
from pathlib import Path

import pyarrow as pa
import pyarrow.parquet as pq

from app.core.config import settings
from app.core.exceptions import AppError
from app.core.logging import get_logger

logger = get_logger(__name__)

EXCEL_BATCH_ROWS = 50_000


def temp_path(suffix: str = ".parquet") -> Path:
    """Reserve a temp filename and close the descriptor.

    ``mkstemp`` hands back an *open* fd. Leaving it open keeps a Windows lock
    on the file, so a later ``os.replace``/``shutil.move`` fails with
    WinError 32.
    """
    fd, path = tempfile.mkstemp(suffix=suffix)
    os.close(fd)
    return Path(path)


class ExcelTooLargeError(AppError):
    code = "EXCEL_TOO_LARGE"
    status_code = 413


def excel_to_parquet(source: Path, destination: Path | None = None) -> Path:
    """Stream an xlsx worksheet into Parquet.

    openpyxl's read-only mode yields rows lazily, and rows are flushed to
    Parquet in batches, so memory stays bounded even though xlsx itself cannot
    be streamed by DuckDB.
    """
    from openpyxl import load_workbook

    destination = destination or temp_path(".parquet")
    workbook = load_workbook(source, read_only=True, data_only=True)
    try:
        sheet = workbook.worksheets[0]
        rows = sheet.iter_rows(values_only=True)

        try:
            header = next(rows)
        except StopIteration as exc:
            raise AppError("The spreadsheet is empty.", code="EMPTY_FILE") from exc

        columns = [
            str(name).strip() if name is not None else f"column_{index}"
            for index, name in enumerate(header)
        ]
        # Deduplicate so Parquet does not reject the schema; the validator
        # still reports the duplication to the user.
        seen: dict[str, int] = {}
        unique_columns: list[str] = []
        for name in columns:
            if name in seen:
                seen[name] += 1
                unique_columns.append(f"{name}_{seen[name]}")
            else:
                seen[name] = 0
                unique_columns.append(name)

        writer: pq.ParquetWriter | None = None
        buffer: list[tuple] = []
        total = 0

        def flush() -> None:
            nonlocal writer, buffer
            if not buffer:
                return
            # Everything is written as text; the profiler's stage-A pass infers
            # real types from the values, exactly as it does for CSV.
            arrays = [
                pa.array([("" if row[i] is None else str(row[i])) if i < len(row) else "" for row in buffer])
                for i in range(len(unique_columns))
            ]
            table = pa.Table.from_arrays(arrays, names=unique_columns)
            if writer is None:
                writer = pq.ParquetWriter(destination, table.schema, compression="zstd")
            writer.write_table(table)
            buffer = []

        try:
            for row in rows:
                if all(value is None for value in row):
                    continue
                buffer.append(row)
                total += 1
                if total > settings.excel_max_rows:
                    raise ExcelTooLargeError(
                        f"The spreadsheet exceeds {settings.excel_max_rows} rows.",
                        details={"limit_rows": settings.excel_max_rows},
                    )
                if len(buffer) >= EXCEL_BATCH_ROWS:
                    flush()
            flush()
            if writer is None:
                # Header but no data rows - emit an empty typed file so the
                # profiler reports NO_ROWS rather than failing to read.
                table = pa.Table.from_arrays(
                    [pa.array([], type=pa.string()) for _ in unique_columns], names=unique_columns
                )
                writer = pq.ParquetWriter(destination, table.schema, compression="zstd")
                writer.write_table(table)
        finally:
            if writer is not None:
                writer.close()
    finally:
        workbook.close()

    return destination


def rows_to_parquet(
    columns: list[str],
    row_batches,  # Iterable[list[tuple]]
    destination: Path,
    *,
    max_rows: int | None = None,
) -> int:
    """Write batches of tuples to Parquet. Used by save-query-as-dataset."""
    writer: pq.ParquetWriter | None = None
    total = 0
    try:
        for batch in row_batches:
            if not batch:
                continue
            arrays = [
                pa.array(
                    [
                        None if index >= len(row) or row[index] is None else str(row[index])
                        for row in batch
                    ]
                )
                for index in range(len(columns))
            ]
            table = pa.Table.from_arrays(arrays, names=columns)
            if writer is None:
                writer = pq.ParquetWriter(destination, table.schema, compression="zstd")
            writer.write_table(table)
            total += len(batch)
            if max_rows is not None and total >= max_rows:
                break
        if writer is None:
            table = pa.Table.from_arrays(
                [pa.array([], type=pa.string()) for _ in columns], names=columns
            )
            writer = pq.ParquetWriter(destination, table.schema, compression="zstd")
            writer.write_table(table)
    finally:
        if writer is not None:
            writer.close()
    return total
